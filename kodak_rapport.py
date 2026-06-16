#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
KODAK — Rapport Statistiques Impressions (GUI)
Génère un fichier Excel avec les stats par mois pour tous les magasins.

Usage GUI:  python kodak_rapport.py
Usage CLI:  python kodak_rapport.py <dossier_compteurs> [sortie.xlsx]
"""

import json, os, sys, glob
from datetime import datetime

def _resource(filename):
    """Chemin vers un fichier ressource (compatible PyInstaller --onefile)."""
    if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, filename)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl requis: pip install openpyxl")
    sys.exit(1)

APP_VERSION = "1.0.2"

MOIS_FR = {
    "01":"Janvier","02":"Février","03":"Mars","04":"Avril",
    "05":"Mai","06":"Juin","07":"Juillet","08":"Août",
    "09":"Septembre","10":"Octobre","11":"Novembre","12":"Décembre"
}

def load_counter(path):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return {}

def magasin_name(path):
    return os.path.splitext(os.path.basename(path))[0]

def collect_data(folder):
    files = glob.glob(os.path.join(folder, "*.json"))
    data = {}
    all_years = set()
    for f in files:
        name = magasin_name(f)
        ctr = load_counter(f)
        data[name] = {}
        for year, ydata in ctr.items():
            if not isinstance(ydata, dict):
                continue
            all_years.add(year)
            data[name][year] = {}
            mois = ydata.get("mois", {})
            for m, mdata in mois.items():
                total = mdata.get("total", 0) if isinstance(mdata, dict) else mdata
                data[name][year][m] = total
    # S'il n'y a aucune année (tous les fichiers vides), mettre l'année courante
    if not all_years:
        all_years.add(str(datetime.now().year))
    return data, sorted(all_years)

def build_excel(data, years, output, mag_names=None):
    wb = Workbook()
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="1F2937")
    title_font = Font(name="Arial", bold=True, size=14, color="D97706")
    sub_font = Font(name="Arial", size=9, color="6B7280")
    data_font = Font(name="Arial", size=11)
    total_font = Font(name="Arial", bold=True, size=11, color="D97706")
    total_fill = PatternFill("solid", fgColor="FEF3C7")
    border = Border(
        left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"))
    center = Alignment(horizontal="center", vertical="center")
    magasins = sorted(data.keys())
    all_months = ["01","02","03","04","05","06","07","08","09","10","11","12"]

    # ==========================================================
    #  ONGLET 1 : DONNÉES BRUTES (source pour TCD)
    # ==========================================================
    ws_data = wb.active
    ws_data.title = "Données"
    headers = ["Magasin", "Année", "Mois", "Mois_Num", "Impressions"]
    for ci, h in enumerate(headers, 1):
        c = ws_data.cell(row=1, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.border = border; c.alignment = center
    ws_data.column_dimensions["A"].width = 40
    ws_data.column_dimensions["B"].width = 10
    ws_data.column_dimensions["C"].width = 14
    ws_data.column_dimensions["D"].width = 12
    ws_data.column_dimensions["E"].width = 14

    row = 2
    for mag in magasins:
        display = mag_names.get(mag, mag) if mag_names else mag
        for year in years:
            mag_year = data.get(mag, {}).get(year, {})
            for m in all_months:
                val = mag_year.get(m, 0)
                ws_data.cell(row=row, column=1, value=display).font = data_font
                ws_data.cell(row=row, column=1).border = border
                ws_data.cell(row=row, column=2, value=int(year)).font = data_font
                ws_data.cell(row=row, column=2).border = border
                ws_data.cell(row=row, column=2).alignment = center
                ws_data.cell(row=row, column=3, value=MOIS_FR[m]).font = data_font
                ws_data.cell(row=row, column=3).border = border
                ws_data.cell(row=row, column=4, value=int(m)).font = data_font
                ws_data.cell(row=row, column=4).border = border
                ws_data.cell(row=row, column=4).alignment = center
                c = ws_data.cell(row=row, column=5, value=val)
                c.font = data_font; c.border = border; c.alignment = center; c.number_format = '#,##0'
                row += 1

    # Appliquer un filtre auto sur les données
    ws_data.auto_filter.ref = f"A1:E{row-1}"

    # ==========================================================
    #  ONGLET PAR ANNÉE : Tableau croisé Magasin x Mois
    # ==========================================================
    for yi, year in enumerate(years):
        ws = wb.create_sheet(title=f"Année {year}")
        ws.merge_cells("A1:N1")
        ws["A1"] = f"Statistiques Impressions — {year}"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 30
        ws.merge_cells("A2:N2")
        ws["A2"] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws["A2"].font = sub_font

        r0 = 4
        c = ws.cell(row=r0, column=1, value="Magasin")
        c.font = hdr_font; c.fill = hdr_fill; c.border = border; c.alignment = center
        ws.column_dimensions["A"].width = 40
        for i, m in enumerate(all_months):
            col = i + 2
            c = ws.cell(row=r0, column=col, value=MOIS_FR[m])
            c.font = hdr_font; c.fill = hdr_fill; c.border = border; c.alignment = center
            ws.column_dimensions[get_column_letter(col)].width = 12
        col_total = 14
        c = ws.cell(row=r0, column=col_total, value="TOTAL")
        c.font = hdr_font; c.fill = hdr_fill; c.border = border; c.alignment = center
        ws.column_dimensions[get_column_letter(col_total)].width = 14

        for mi, mag in enumerate(magasins):
            r = r0 + 1 + mi
            display = mag_names.get(mag, mag) if mag_names else mag
            ws.cell(row=r, column=1, value=display).font = data_font
            ws.cell(row=r, column=1).border = border
            mag_year = data.get(mag, {}).get(year, {})
            for i, m in enumerate(all_months):
                col = i + 2
                c = ws.cell(row=r, column=col, value=mag_year.get(m, 0))
                c.font = data_font; c.border = border; c.alignment = center; c.number_format = '#,##0'
            c = ws.cell(row=r, column=col_total)
            c.value = f"=SUM(B{r}:M{r})"
            c.font = total_font; c.fill = total_fill; c.border = border; c.alignment = center; c.number_format = '#,##0'

        # Ligne TOTAL
        r_total = r0 + 1 + len(magasins)
        c = ws.cell(row=r_total, column=1, value="TOTAL")
        c.font = total_font; c.fill = total_fill; c.border = border
        for col in range(2, col_total + 1):
            cl = get_column_letter(col)
            c = ws.cell(row=r_total, column=col)
            c.value = f"=SUM({cl}{r0+1}:{cl}{r0+len(magasins)})"
            c.font = total_font; c.fill = total_fill; c.border = border; c.alignment = center; c.number_format = '#,##0'

        # Graphique
        if magasins:
            chart = BarChart()
            chart.type = "col"; chart.style = 10
            chart.title = f"Impressions par mois — {year}"
            chart.y_axis.title = "Impressions"; chart.x_axis.title = "Mois"
            chart.width = 28; chart.height = 14
            cats = Reference(ws, min_col=2, max_col=13, min_row=r0)
            for mi, mag in enumerate(magasins):
                r = r0 + 1 + mi
                vals = Reference(ws, min_col=2, max_col=13, min_row=r)
                chart.add_data(vals, from_rows=True, titles_from_data=False)
                display = mag_names.get(mag, mag) if mag_names else mag
                chart.series[mi].tx = SeriesLabel(v=display)
            chart.set_categories(cats)
            ws.add_chart(chart, f"A{r_total + 3}")

    # ==========================================================
    #  ONGLET SYNTHÈSE MULTI-ANNÉES
    # ==========================================================
    if len(years) >= 1:
        ws_syn = wb.create_sheet(title="Synthèse")
        ws_syn.merge_cells(f"A1:{get_column_letter(len(years)+2)}1")
        ws_syn["A1"] = "Synthèse par magasin et par année"
        ws_syn["A1"].font = title_font
        ws_syn["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws_syn.row_dimensions[1].height = 30
        ws_syn.merge_cells(f"A2:{get_column_letter(len(years)+2)}2")
        ws_syn["A2"] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws_syn["A2"].font = sub_font

        r0 = 4
        c = ws_syn.cell(row=r0, column=1, value="Magasin")
        c.font = hdr_font; c.fill = hdr_fill; c.border = border; c.alignment = center
        ws_syn.column_dimensions["A"].width = 40
        for yi, year in enumerate(years):
            col = yi + 2
            c = ws_syn.cell(row=r0, column=col, value=year)
            c.font = hdr_font; c.fill = hdr_fill; c.border = border; c.alignment = center
            ws_syn.column_dimensions[get_column_letter(col)].width = 14
        col_total = len(years) + 2
        c = ws_syn.cell(row=r0, column=col_total, value="TOTAL")
        c.font = hdr_font; c.fill = hdr_fill; c.border = border; c.alignment = center
        ws_syn.column_dimensions[get_column_letter(col_total)].width = 14

        for mi, mag in enumerate(magasins):
            r = r0 + 1 + mi
            display = mag_names.get(mag, mag) if mag_names else mag
            ws_syn.cell(row=r, column=1, value=display).font = data_font
            ws_syn.cell(row=r, column=1).border = border
            for yi, year in enumerate(years):
                col = yi + 2
                mag_year = data.get(mag, {}).get(year, {})
                total_y = sum(mag_year.get(m, 0) for m in all_months)
                c = ws_syn.cell(row=r, column=col, value=total_y)
                c.font = data_font; c.border = border; c.alignment = center; c.number_format = '#,##0'
            c = ws_syn.cell(row=r, column=col_total)
            c.value = f"=SUM({get_column_letter(2)}{r}:{get_column_letter(len(years)+1)}{r})"
            c.font = total_font; c.fill = total_fill; c.border = border; c.alignment = center; c.number_format = '#,##0'

        # Ligne TOTAL synthèse
        r_total = r0 + 1 + len(magasins)
        c = ws_syn.cell(row=r_total, column=1, value="TOTAL")
        c.font = total_font; c.fill = total_fill; c.border = border
        for col in range(2, col_total + 1):
            cl = get_column_letter(col)
            c = ws_syn.cell(row=r_total, column=col)
            c.value = f"=SUM({cl}{r0+1}:{cl}{r0+len(magasins)})"
            c.font = total_font; c.fill = total_fill; c.border = border; c.alignment = center; c.number_format = '#,##0'

    wb.save(output)


# ============================================================================
#  GUI
# ============================================================================

def run_gui():
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox

    BG="#0d1117"; BC="#161b22"; BD="#010409"; BO="#30363d"
    TX="#e6edf3"; TD="#7d8590"; AM="#f0883e"; GR="#3fb950"; RD="#f85149"; BL="#58a6ff"

    root = tk.Tk()
    root.title("Kodak — Rapport Impressions")
    try:
        root.iconbitmap(_resource("kodak_rapport.ico"))
    except Exception:
        try:
            _app_icon = tk.PhotoImage(file=_resource("icone-imprimante.png"))
            root.iconphoto(True, _app_icon)
        except Exception:
            pass
    root.configure(bg=BG)
    root.geometry("820x640")
    root.minsize(700, 500)

    # Header
    hdr = tk.Frame(root, bg="#0b0e14", padx=20, pady=12); hdr.pack(fill=tk.X)
    tk.Label(hdr, text=" K ", font=("Consolas", 16, "bold"), bg=AM, fg="#fff", padx=6).pack(side=tk.LEFT)
    tf = tk.Frame(hdr, bg="#0b0e14"); tf.pack(side=tk.LEFT, padx=(12, 0))
    tk.Label(tf, text="RAPPORT STATISTIQUES", font=("Segoe UI", 13, "bold"), bg="#0b0e14", fg=TX).pack(anchor="w")
    tk.Label(tf, text="Impressions par mois — Multi-magasins", font=("Segoe UI", 9), bg="#0b0e14", fg=TD).pack(anchor="w")
    tk.Label(tf, text=f"v{APP_VERSION}", font=("Segoe UI", 8), bg="#0b0e14", fg="#3a4a5a").pack(anchor="w")

    # Dossier source
    src_frame = tk.Frame(root, bg=BC, padx=20, pady=14, highlightbackground=BO, highlightthickness=1)
    src_frame.pack(fill=tk.X, padx=20, pady=(14, 0))
    tk.Label(src_frame, text="DOSSIER DES COMPTEURS", font=("Segoe UI", 11, "bold"), bg=BC, fg=AM).pack(anchor="w", pady=(0, 8))
    row_src = tk.Frame(src_frame, bg=BC); row_src.pack(fill=tk.X)
    tk.Label(row_src, text="Dossier:", font=("Segoe UI", 10), bg=BC, fg=TD, width=10, anchor="w").pack(side=tk.LEFT)
    src_var = tk.StringVar()
    tk.Entry(row_src, textvariable=src_var, font=("Consolas", 10), bg=BD, fg=TX,
             insertbackground=TX, relief="flat", highlightthickness=1, highlightbackground=BO, width=50).pack(side=tk.LEFT, padx=(4, 4))
    def browse_src():
        d = filedialog.askdirectory(title="Dossier des compteurs")
        if d:
            src_var.set(d)
            refresh_list()
    tk.Button(row_src, text="...", command=browse_src, bg="#1c2333", fg=TX, relief="flat", padx=8).pack(side=tk.LEFT)
    tk.Button(row_src, text="↻", command=lambda: refresh_list(), bg="#1c2333", fg=BL,
              relief="flat", padx=8, font=("Segoe UI", 11, "bold")).pack(side=tk.LEFT, padx=(4, 0))

    # Liste magasins
    list_frame = tk.Frame(root, bg=BC, padx=20, pady=14, highlightbackground=BO, highlightthickness=1)
    list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(14, 0))
    lbl_row = tk.Frame(list_frame, bg=BC); lbl_row.pack(fill=tk.X, pady=(0, 8))
    tk.Label(lbl_row, text="MAGASINS DÉTECTÉS", font=("Segoe UI", 11, "bold"), bg=BC, fg=AM).pack(side=tk.LEFT)
    tk.Label(lbl_row, text="Double-clic sur le nom pour renommer", font=("Segoe UI", 9), bg=BC, fg=TD).pack(side=tk.RIGHT)

    sty = ttk.Style(); sty.theme_use("clam")
    sty.configure("M.Treeview", background=BC, foreground=TX, fieldbackground=BC, font=("Segoe UI", 10), rowheight=28)
    sty.configure("M.Treeview.Heading", background="#1c2333", foreground=AM, font=("Segoe UI", 10, "bold"))
    sty.map("M.Treeview", background=[("selected", "#1c3a5e")])

    tree_frame = tk.Frame(list_frame, bg=BC); tree_frame.pack(fill=tk.BOTH, expand=True)
    tree = ttk.Treeview(tree_frame, columns=("file", "name", "total"), show="headings", style="M.Treeview")
    tree.heading("file", text="Fichier"); tree.heading("name", text="Nom affiché"); tree.heading("total", text="Total")
    tree.column("file", width=220); tree.column("name", width=250); tree.column("total", width=100, anchor="center")
    tsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=tsb.set)
    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True); tsb.pack(side=tk.RIGHT, fill=tk.Y)

    mag_names = {}

    def on_double_click(event):
        item = tree.focus()
        if not item: return
        col = tree.identify_column(event.x)
        if col != "#2": return
        bbox = tree.bbox(item, "name")
        if not bbox: return
        x, y, w, h = bbox
        entry = tk.Entry(tree, font=("Segoe UI", 10), bg=BD, fg=TX, insertbackground=TX, relief="flat")
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, mag_names.get(item, item))
        entry.select_range(0, tk.END)
        entry.focus()
        def save(e=None):
            v = entry.get().strip()
            if v:
                mag_names[item] = v
                tree.set(item, "name", v)
            entry.destroy()
        entry.bind("<Return>", save)
        entry.bind("<FocusOut>", save)
        entry.bind("<Escape>", lambda e: entry.destroy())
    tree.bind("<Double-1>", on_double_click)

    # Sortie
    out_frame = tk.Frame(root, bg=BC, padx=20, pady=14, highlightbackground=BO, highlightthickness=1)
    out_frame.pack(fill=tk.X, padx=20, pady=(14, 0))
    row_out = tk.Frame(out_frame, bg=BC); row_out.pack(fill=tk.X)
    tk.Label(row_out, text="Fichier sortie:", font=("Segoe UI", 10), bg=BC, fg=TD, width=12, anchor="w").pack(side=tk.LEFT)
    out_var = tk.StringVar(value=f"rapport_impressions_{datetime.now().strftime('%Y%m')}.xlsx")
    tk.Entry(row_out, textvariable=out_var, font=("Consolas", 10), bg=BD, fg=TX,
             insertbackground=TX, relief="flat", highlightthickness=1, highlightbackground=BO, width=40).pack(side=tk.LEFT, padx=(4, 4))
    def browse_out():
        f = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
                                          initialfile=out_var.get(), title="Enregistrer le rapport")
        if f: out_var.set(f)
    tk.Button(row_out, text="...", command=browse_out, bg="#1c2333", fg=TX, relief="flat", padx=8).pack(side=tk.LEFT)

    # Footer
    foot = tk.Frame(root, bg=BG, padx=20, pady=12); foot.pack(fill=tk.X)
    voyant_dot = tk.Label(foot, text="●", font=("Segoe UI", 11), bg=BG, fg=TD)
    voyant_dot.pack(side=tk.LEFT, padx=(0, 6))
    status_lbl = tk.Label(foot, text="Sélectionnez un dossier", font=("Segoe UI", 9), bg=BG, fg=TD)
    status_lbl.pack(side=tk.LEFT)
    tk.Label(foot, text=f"v{APP_VERSION}", font=("Segoe UI", 8), bg=BG, fg="#2a3a4a").pack(side=tk.RIGHT)

    def set_voyant(color, text):
        voyant_dot.config(fg=color)
        status_lbl.config(text=text, fg=color)

    def refresh_list():
        for i in tree.get_children(): tree.delete(i)
        mag_names.clear()
        folder = src_var.get().strip()
        if not folder or not os.path.isdir(folder):
            set_voyant(TD, "Sélectionnez un dossier"); return
        files = sorted(glob.glob(os.path.join(folder, "*.json")))
        for f in files:
            base = magasin_name(f)
            ctr = load_counter(f)
            total = sum(y.get("total", 0) for y in ctr.values() if isinstance(y, dict))
            mag_names[base] = base
            tree.insert("", "end", iid=base, values=(base, base, f"{total:,}"))
        if mag_names:
            set_voyant("#58a6ff", f"{len(mag_names)} magasin(s) — Prêt à générer")
        else:
            set_voyant(TD, "Aucun fichier compteur dans ce dossier")

    def generate():
        folder = src_var.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showerror("Erreur", "Sélectionnez un dossier valide."); return
        if not mag_names:
            messagebox.showerror("Erreur", "Aucun fichier compteur trouvé."); return
        out = out_var.get().strip()
        if not out:
            messagebox.showerror("Erreur", "Spécifiez un fichier de sortie."); return
        if not os.path.isabs(out):
            out = os.path.join(folder, out)
        gen_btn.config(state="disabled", text="Génération...")
        set_voyant("#d29922", "Génération en cours...")
        root.update()
        try:
            data, years = collect_data(folder)
            build_excel(data, years, out, mag_names)
            set_voyant(GR, f"Rapport généré : {os.path.basename(out)}")
            messagebox.showinfo("Succès", f"Rapport généré :\n{out}")
        except Exception as e:
            set_voyant(RD, f"Erreur : {e}")
            messagebox.showerror("Erreur", str(e))
        finally:
            gen_btn.config(state="normal", text="Générer le rapport")

    gen_btn = tk.Button(foot, text="Générer le rapport", command=generate,
                        bg="#1a4d2e", fg=GR, font=("Segoe UI", 12, "bold"),
                        relief="flat", padx=20, pady=6, cursor="hand2")
    gen_btn.pack(side=tk.RIGHT)
    root.mainloop()


if __name__ == "__main__":
    if len(sys.argv) > 1:
        folder = sys.argv[1]
        if not os.path.isdir(folder):
            print(f"Dossier introuvable: {folder}"); sys.exit(1)
        output = sys.argv[2] if len(sys.argv) > 2 else os.path.join(folder, f"rapport_impressions_{datetime.now().strftime('%Y%m')}.xlsx")
        data, years = collect_data(folder)
        print(f"{len(data)} magasin(s): {', '.join(sorted(data.keys()))}")
        build_excel(data, years, output)
        print(f"Rapport: {output}")
    else:
        run_gui()
